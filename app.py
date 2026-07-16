"""
PlayMaker Pro - dashboard skautingowy (prototyp).
Uklad: filtry -> karty topowych (scroll w bok) -> tabela graczy -> klikalne mecze.

Uruchomienie:
    pip install -r requirements.txt
    streamlit run app.py   # pliki: stats_test.csv, matches_test.csv, teamy_kluby_25_26.csv
"""
import os
import io
import re
import json
import numpy as np
import pandas as pd
import streamlit as st

def _secret(key, default=""):
    """Czyta najpierw st.secrets (Streamlit Cloud), potem zmienne środowiskowe."""
    try:
        if key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass
    return os.environ.get(key, default)


DATA_MODE = _secret("PM_DATA_MODE", "csv") or "csv"
REGION = _secret("PM_REGION", "")   # region wybrany na wejsciu (region -> liga -> play)

# Rodzaj gramatyczny etykiet UI: 'm' (domyślnie, kluby) albo 'f' (raport dziewczynek).
# Ustaw sekret PM_GENDER = "f", żeby przełączyć teksty na żeńskie.
_GENDER_F = (_secret("PM_GENDER", "m") or "m").lower().startswith("f")
L = {
    "player_one":  "Piłkarka"               if _GENDER_F else "Zawodnik",
    "players_gen": "zawodniczek"            if _GENDER_F else "zawodników",
    "players_all": "Wszystkie zawodniczki"  if _GENDER_F else "Wszyscy zawodnicy",
    "of_player":   "zawodniczki"            if _GENDER_F else "zawodnika",
    "played":      "grała"                  if _GENDER_F else "grał",
    "top_players": "Topowe piłkarki"        if _GENDER_F else "Topowi zawodnicy",
    "no_players":  "Brak zawodniczek dla wybranych filtrów." if _GENDER_F
                   else "Brak zawodników dla wybranych filtrów.",
    "click_one":   "kliknij piłkarkę"       if _GENDER_F else "kliknij gracza",
}


NUMERIC_COMMA = ["match_score", "m_overall_score", "m_season_score", "overall_score",
                 "season_score", "global_last_overall_score", "global_last_season_score"]
NUMERIC_PLAIN = ["minutes", "goals", "yellow_cards", "red_cards", "est_birth_year",
                 "age_at_match", "senior_minutes", "senior_matches_played",
                 "senior_squad_apps", "play_cohort_birth_year", "matches_count"]
PM_WEIGHTS = {"Jakość": 0.55, "Forma": 0.10, "Dostępność": 0.20, "Konsekwencja": 0.15}

# ===== PlayMaker Score 2.0 — port produkcyjnego wzoru (ścieżka domyślna, bez pozycji) =====
# Źródła: DefaultMatchScoreStrategy.ts, ScoreConstants.ts, AgeScoreCalculator.ts,
# DefaultSegmentFactorCalculator.ts. Dla młodzieży leagueMultiplier bierzemy z rank_p (v7),
# bo produkcyjnie KAŻDA kategoria (A1/B1/C1...) ma płaski multiplier ~0.08 niezależnie od poziomu.
PM_SCORE_MODE = _secret("PM_SCORE_MODE", "v7")   # 'v7' (realny score) lub 'prod' (stary match_score)
# Zakres rankowania jakości: 'selected' (domyślnie — w obrębie wybranej ligi, dla klubów)
# albo 'total' (ze WSZYSTKICH meczów sezonu — dla raportów "grali w starszych/w górę").
QUALITY_SCOPE = (_secret("PM_QUALITY_SCOPE", "selected") or "selected").lower()

# Tryb rankingu: 'standard' (PM Index jak dotąd) albo 'talent' (mix dla raportów "grali w górę":
# jakość league-aware + poziom [CLJ/seniorzy/skok 2+ roczniki] + wolumen). Wagi strojone sekretami.
PM_RANK_MODE = (_secret("PM_RANK_MODE", "standard") or "standard").lower()
def _fnum(key, dflt):
    try:
        v = _secret(key, "")
        return float(v) if v not in ("", None) else float(dflt)
    except Exception:
        return float(dflt)
W_JAKOSC = _fnum("PM_W_JAKOSC", 0.45)    # waga jakości (league-aware)
W_POZIOM = _fnum("PM_W_POZIOM", 0.45)    # waga poziomu (CLJ/seniorzy/duży skok)
W_WOLUMEN = _fnum("PM_W_WOLUMEN", 0.10)  # waga wolumenu (minuty łącznie)
W_CLJ = _fnum("PM_W_CLJ", 1.0)           # w "poziomie": waga minut CLJ
W_SENIOR = _fnum("PM_W_SENIOR", 0.8)     # w "poziomie": waga minut w seniorach
W_SKOK = _fnum("PM_W_SKOK", 0.6)         # w "poziomie": waga minut 2+ roczniki w górę

AGE_IMPACT = 0.35
AGE_CONST = 0.816496580927726
OPTIMAL_AGE = 26
SEGMENT_CONST = 0.947368421052632
SEGMENT_IMPACT_DEFAULT = 0.8
STATS_IMPACT = 0.65
MATCH_IMPACT_DEFAULT = 0.2
MAX_MATCH_TIME = 90
LVL_DECAY = 0.82
_SEG_RATIO = {"wygrana": 1.0, "remis": 0.65, "porażka": 0.4, "porazka": 0.4}
_RESULT_RATIO = {"wygrana": 1.0, "remis": 0.65, "porażka": 0.4, "porazka": 0.4}
_SIDE_RATIO = {"gospodarz": 0.74212204820255, "gość": 1.0, "gosc": 1.0}
# rank_l: 1=CLJ U-19, 2=Makroregionalna, 3=A1, 4=A2, 5=CLJ U-17, 6=B1, 7=B2,
#         8=CLJ U-15, 9=C1, 10=C2, 11=D1, 12=D2
# UWAGA: kolejność odzwierciedla JAKOŚĆ rozgrywek, nie wiek kategorii. CLJ to rozgrywki
# krajowe i stoją wyżej niż regionalne A1/A2, mimo że A1 jest "starsza".
# (przed korektą CLJ U-17 miało 0.20 < A1 0.25 — wiejska A1 biła CLJ U-17 w Lechu)
_BASE_RATIO = {1: 0.38, 2: 0.31, 3: 0.25, 4: 0.21, 5: 0.34, 6: 0.17,
               7: 0.15, 8: 0.30, 9: 0.14, 10: 0.13, 11: 0.12, 12: 0.10}
_EXPECTED_AGE = {1: 19, 2: 19, 3: 19, 4: 18, 5: 17, 6: 17, 7: 16,
                 8: 15, 9: 15, 10: 14, 11: 13, 12: 12}
# Wiek, od którego gra w SENIORACH jest "normalna". Młodszy zawodnik w seniorach gra
# w górę i należy mu się ten sam mechanizm ageDiscount co juniorowi w starszej kategorii.
# Bez tego rank_l był pusty dla seniorów -> exp_age NaN -> dd NaN -> disc twardo 1.00,
# czyli 15-latek w 2. lidze nie dostawał nic za najtrudniejszą rzecz, jaką może zrobić.
SENIOR_EXPECTED_AGE = 19

# Skok wiekowy (ageDiscount) liczy się PROPORCJONALNIE do siły rozgrywek, do których
# zawodnik skacze. Gra 5 lat w górę w wiejskiej III lidze okręgowej A1 to zwykle brak
# odpowiedniej drużyny w klubie, a nie talent — i nie może bić CLJ U-15 w akademii.
# Przy leagueMultiplier >= SKOK_LM_REF premia działa w pełni (CLJ, ligi seniorskie).
SKOK_LM_REF = 0.30
SENIOR_LR = {
    "337bb869-0b42-484f-8eca-0c8842a13ec9": 1.0,    # Ekstraklasa
    "50e40483-e8dc-4e4b-9f58-a83f93a54d9a": 0.9,    # 1 liga
    "5f26d625-e72e-4aa5-9ffe-451025c18e3a": 0.8,    # 2 liga
    "5cc45e5f-744b-428c-b8af-cdefca38de29": 0.65,   # 3 liga
    "c164ca31-22e4-43fc-9e30-4f3bcc2b7d72": 0.4,    # 4 liga (generic)
    "a0583713-115c-4aa5-90f2-140f6eaece15": 0.23,   # 5 liga
    "c5afdf4b-b449-4ef3-acf5-dded47fc5f58": 0.2,    # klasa okręgowa
    "63d04023-727a-4c0c-a8c6-4154fe1104b7": 0.12,   # klasa A
    "b7d2c55b-e2af-44e2-9df2-3f6e05dc1768": 0.1,    # klasa B
    "895016b3-4fa6-4a68-aa41-5035f9ebef8e": 0.1,    # klasa C
}
YOUTH_LID_RANK_L = {
    "bf74d613-4cc6-4115-ad03-fac139dee351": 1, "8e70e715-3f0f-4481-a01d-51fb7b9aee90": 2,
    "5b788871-3d38-4073-9500-fcfa4d1b4270": 3, "f19d92f4-14f7-45ab-884f-90da0d03f4a0": 4,
    "8104ee44-740c-4f6c-8fc3-3bbcf2b3b0e7": 5, "823a45df-052b-4cd5-a060-32ed52921992": 6,
    "75b51f36-93fd-49cb-86d3-6086dc88081b": 7, "436dc4c6-bc94-4d30-ae92-1113d6d4eee3": 8,
    "317d1eb3-4873-4749-91b5-edb2d0cd4375": 9, "adf4ca7f-46ef-4aff-a7a0-3e7cc614c59d": 10,
}
_RANK_L_NAME = [(r"clj u-?19|junior[óo]w u-?19", 1), (r"makroregional", 2), (r"\ba1\b", 3),
                (r"\ba2\b", 4), (r"clj u-?17|junior[óo]w u-?17", 5), (r"\bb1\b", 6),
                (r"\bb2\b", 7), (r"clj u-?15|junior[óo]w u-?15|juniorek", 8), (r"\bc1\b", 9),
                (r"\bc2\b", 10), (r"\bd1\b", 11), (r"\bd2\b", 12)]

try:
    with open("play_id_rank_p_v7.json", encoding="utf-8") as _f:
        RANK_P_DICT = {k: int(v) for k, v in json.load(_f).items()}
except Exception:
    RANK_P_DICT = {}


def _detect_rank_p(name: str) -> int:
    """Fallback poziomu rozgrywek po nazwie play (zgodny z junior_league_resolver v7)."""
    n = name.lower()
    has_city = bool(re.match(r"^[a-ząćęłńóśźż\s\-\.]+:", name, re.IGNORECASE))
    if re.search(r"\bclj\b|centralna liga junior", n):
        return 0
    if re.search(r"liga makroregionalna", n) and not has_city:
        return 0
    if not has_city and re.search(r"\bi liga wojewódzka\b", n):
        return 2
    if not has_city and re.search(r"\bii liga wojewódzka\b", n):
        return 3
    if not has_city and re.search(r"\biii liga wojewódzka\b|\biv liga wojewódzka\b", n):
        return 4
    if re.search(r"\bliga wojewódzka\b", n):
        return 2
    if re.search(r"\bi liga okręgowa\b|\bklasa okręgowa\b", n):
        return 5
    if re.search(r"\bii liga okręgowa\b", n):
        return 6
    if re.search(r"\biii liga okręgowa\b", n):
        return 7
    if re.search(r"\biv liga okręgowa\b", n):
        return 8
    if re.search(r"\bv liga okręgowa\b|\bvi liga okręgowa\b|\bvii liga okręgowa\b", n):
        return 9
    return 10


def _rank_l_name(name):
    n = str(name).lower()
    for pat, rl in _RANK_L_NAME:
        if re.search(pat, n):
            return rl
    return np.nan


def _rank_l_series(df):
    rl = df["league_id"].astype(str).map(YOUTH_LID_RANK_L) if "league_id" in df.columns else None
    if rl is None:
        rl = pd.Series(np.nan, index=df.index)
    miss = rl.isna()
    if miss.any():
        uniq = {u: _rank_l_name(u) for u in df.loc[miss, "league_name"].astype(str).unique()}
        rl = rl.where(~miss, df.loc[miss, "league_name"].astype(str).map(uniq))
    return pd.to_numeric(rl, errors="coerce")


def _rank_p_series(df):
    rp = df["play_id"].map(RANK_P_DICT)
    miss = rp.isna()
    if miss.any():
        uniq = {u: _detect_rank_p(u) for u in df.loc[miss, "play_name"].astype(str).unique()}
        rp = rp.where(~miss, df.loc[miss, "play_name"].astype(str).map(uniq))
    return pd.to_numeric(rp, errors="coerce")


# Sezon 25/26 (PZPN): każda dywizja ma JEDEN najstarszy dopuszczalny rocznik (max wiek).
# Młodszy może grać w każdej starszej; "za stary" nie zagra niżej. Gra w dywizji o
# starszym max-roczniku niż własny = "gra ze starszymi".
_CAT_MAXYEAR_PATS = [
    (r'(^A1$|U-?19)', 2007), (r'(^A2$|U-?18)', 2008),
    (r'(^B1$|U-?17)', 2009), (r'(^B2$|U-?16)', 2010),
    (r'(^C1$|U-?15)', 2011), (r'(^C2$|U-?14)', 2012),
    (r'(^D1$|U-?13)', 2013), (r'(^D2$|U-?12)', 2014),
    (r'(^E1$|U-?11)', 2015), (r'(^E2$|U-?10)', 2016),
    (r'(^F1$|U-?9)', 2017),  (r'(^F2$|U-?8)', 2018),
]


def _cat_max_year(name):
    """Najstarszy dopuszczalny rocznik dla dywizji (sezon 25/26). CLJ U-1x wg numeru.
    Senior / nieznane juniorskie -> NaN (nie liczymy jako 'w górę')."""
    n = str(name)
    for pat, y in _CAT_MAXYEAR_PATS:
        if re.search(pat, n, re.I):
            return y
    return np.nan


def _cat_maxyear_series(df):
    return df["league_name"].map(_cat_max_year)


def _rozgrywki_key(name):
    """Kanoniczna nazwa rozgrywek — scala rundy jesień/wiosna, baraże i (RW) tej samej ligi
    (źródło rozbija je na osobne play_id)."""
    n = str(name)
    n = re.sub(r"\s*-?\s*RUNDA\s+\w+", " ", n, flags=re.I)
    n = re.sub(r"bara[żz]\w*.*", " ", n, flags=re.I)
    n = re.sub(r"\(RW\)|\[[^\]]*\]|Grupa\s+\S+", " ", n, flags=re.I)
    n = n.replace('"', " ")
    n = re.sub(r"\s{2,}", " ", n).strip(" -")
    return n or str(name)


_JR_ORDER = {"A1": 0, "A2": 1, "B1": 2, "B2": 3, "C1": 4, "C2": 5,
             "D1": 6, "D2": 7, "E1": 8, "E2": 9, "F1": 10, "F2": 11}


def _liga_rank(league_name, play_name=""):
    """Priorytet wyświetlania lig: seniorzy -> CLJ -> juniorskie wojewódzkie
    (starsze->młodsze) -> okręgowe/niższe. Zwraca krotkę do sortowania."""
    ln = str(league_name).strip().upper()
    s = f"{league_name} {play_name}".lower()
    if re.search(r"clj|centralna liga", s):
        return (1, 0)
    if ln in _JR_ORDER or re.search(r"junior|trampkarz|m[lł]odzik|orlik|[zż]ak|skrzat|u-1", s):
        base = 3 if re.search(r"okręg|klasa", s) else 2
        return (base, _JR_ORDER.get(ln, 5))
    return (0, 0)   # seniorzy (ligi/klasy dorosłych) — najwyżej


def _zaproszenie_pdf(zawodnik, rocznik="", klub_zaw=""):
    """Zwraca bajty PDF 'Zaproszenie na testy' z logo klubu, albo None jeśli brak reportlab."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
    except Exception:
        return None
    import datetime as _dt
    import textwrap as _tw
    import glob as _glob
    klub = _secret("PM_KLUB", "OKS Odra Opole")
    adres = _secret("PM_KLUB_ADRES", "ul. Leonarda Olejnika 1, 45-839 Opole")
    logo = _secret("PM_KLUB_LOGO", "")
    if not (logo and os.path.exists(logo)):
        cand = _glob.glob("logo.*") or _glob.glob("*logo*.*")
        logo = cand[0] if cand else None
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    y = H - 30 * mm
    if logo and os.path.exists(logo):
        try:
            img = ImageReader(logo); iw, ih = img.getSize()
            w = 38 * mm; h = w * ih / iw
            c.drawImage(img, (W - w) / 2, y - h + 10 * mm, width=w, height=h,
                        preserveAspectRatio=True, mask="auto")
            y -= h
        except Exception:
            pass
    c.setFont("Helvetica-Bold", 16); c.drawCentredString(W / 2, y, klub); y -= 7 * mm
    c.setFont("Helvetica", 9); c.drawCentredString(W / 2, y, adres); y -= 6 * mm
    c.setStrokeColor(colors.HexColor("#c0392b")); c.setLineWidth(1.2)
    c.line(25 * mm, y, W - 25 * mm, y); y -= 18 * mm
    c.setFont("Helvetica", 10)
    c.drawRightString(W - 25 * mm, y, "Opole, dnia " + _dt.date.today().strftime("%d.%m.%Y") + " r."); y -= 16 * mm
    c.setFont("Helvetica-Bold", 15); c.drawCentredString(W / 2, y, "ZAPROSZENIE NA TESTY"); y -= 14 * mm
    c.setFont("Helvetica", 11); t = c.beginText(25 * mm, y); t.setLeading(16)
    t.textLine("Szanowni Państwo,"); t.textLine("")
    linia = (f"{klub} ma przyjemność zaprosić zawodnika {zawodnik}"
             + (f" (rocznik {rocznik})" if rocznik else "")
             + (f", {klub_zaw}," if klub_zaw else "")
             + " na testy do drużyny młodzieżowej naszego klubu.")
    for w in _tw.wrap(linia, 92):
        t.textLine(w)
    t.textLine("")
    for w in _tw.wrap("Prosimy o potwierdzenie obecności oraz zabranie stroju treningowego, "
                      "obuwia na nawierzchnię naturalną i sztuczną oraz ochraniaczy.", 92):
        t.textLine(w)
    c.drawText(t)
    c.setFont("Helvetica", 11)
    c.drawRightString(W - 25 * mm, 45 * mm, "Z wyrazami szacunku,")
    c.drawRightString(W - 25 * mm, 38 * mm, klub)
    c.showPage(); c.save(); buf.seek(0)
    return buf.getvalue()


def compute_pm_score(df):
    """Realny PlayMaker Score 2.0 per mecz (ścieżka domyślna). Zwraca pd.Series w skali 0..1."""
    idx = df.index
    rank_l = _rank_l_series(df)
    rank_p = _rank_p_series(df).fillna(8)
    is_youth = rank_l.notna()
    exp_age = rank_l.map(_EXPECTED_AGE)

    age = pd.to_numeric(df.get("age_at_match"), errors="coerce")
    age = age.fillna(exp_age).fillna(OPTIMAL_AGE)

    # leagueMultiplier: młodzież z rank_p (v7), reszta z produkcyjnej mapy seniorskiej
    # (liczony PRZED agePart, bo waży premię za skok wiekowy — patrz SKOK_LM_REF)
    youth_lm = rank_l.map(_BASE_RATIO) * (LVL_DECAY ** rank_p)
    senior_lm = (df["league_id"].astype(str).map(SENIOR_LR).fillna(0.4)
                 if "league_id" in df.columns else pd.Series(0.4, index=idx))
    lm = youth_lm.where(is_youth, senior_lm)

    # agePart = AGE_IMPACT * AGE_CONST * sqrt(max(0, 1-(age-26)/26)) * ageDiscount(v7)
    normalized = (1 - (age - OPTIMAL_AGE) / OPTIMAL_AGE).clip(lower=0)
    age_base = AGE_IMPACT * AGE_CONST * np.sqrt(normalized)
    # seniorzy też mają odniesienie wiekowe (patrz SENIOR_EXPECTED_AGE)
    exp_age_disc = exp_age.where(is_youth, SENIOR_EXPECTED_AGE)
    dd = (exp_age_disc - age)
    disc = pd.Series(np.select(
        [dd >= 8, dd == 7, dd == 6, dd == 5, dd == 4, dd == 3, dd == 2, dd == 1],
        [1.70, 1.60, 1.50, 1.40, 1.30, 1.20, 1.10, 1.05], default=1.0), index=idx)
    # premia za skok wiekowy proporcjonalna do siły ligi, do której zawodnik skacze
    disc = 1.0 + (disc - 1.0) * (lm / SKOK_LM_REF).clip(upper=1.0)
    age_part = age_base * disc

    res = df["match_result"].astype(str).str.lower()
    mn = pd.to_numeric(df["minutes"], errors="coerce").fillna(0)
    mr = np.where(mn == 0, 0, np.minimum(np.round(mn / 10) * 10, MAX_MATCH_TIME))
    mr = pd.Series(mr, index=idx)
    goals = pd.to_numeric(df["goals"], errors="coerce").fillna(0)
    red = pd.to_numeric(df["red_cards"], errors="coerce").fillna(0)
    seg_ratio = res.map(_SEG_RATIO).fillna(0.4)
    with np.errstate(divide="ignore", invalid="ignore"):
        seg = SEGMENT_CONST * ((goals / mr - red / mr + mr / MAX_MATCH_TIME) * seg_ratio)
    seg = seg.where(mr > 0, 0.0)

    result_ratio = res.map(_RESULT_RATIO).fillna(0.4)
    side_ratio = df["team_side"].astype(str).str.lower().map(_SIDE_RATIO).fillna(1.0)
    match_factor = MATCH_IMPACT_DEFAULT * (result_ratio * side_ratio)
    stats_part = STATS_IMPACT * ((SEGMENT_IMPACT_DEFAULT * seg + match_factor) * lm)
    score = (age_part + stats_part).clip(upper=1.0)
    return pd.DataFrame({"score": score, "stats_part": stats_part}, index=idx)
# premie kontekstowe doliczane do bazy PM Index (tunable)
B_UP = 0.05          # gra ze starszymi rocznikiem (junior w gore)
B_SEN_SQUAD = 0.04   # w kadrze seniorow (0 minut)
B_SEN_PLAYED = 0.12  # realne minuty w seniorach
PM_HELP = (
    "**PM Index** to wskaźnik pozycji zawodnika względem stawki tej ligi — im wyżej, tym "
    "lepiej wypada na tle rywali. Baza (zwykle 0–1) to ważona suma czterech percentyli "
    "liczonych w obrębie wybranej ligi:\n\n"
    "- **Jakość (55%)** — część **league + performance** realnego PlayMaker Score 2.0 "
    "(stats_part: leagueMultiplier × wynik × występ), z leagueMultiplier dla młodzieży z poziomu "
    "rozgrywek (rank_p, v7). Pełny score (z komponentem wieku) jest w kolumnie „PM Score”, ale dla "
    "młodzieży dominuje go wiek, więc ranking bierze stats_part, a sygnał wieku/gry w górę jest w premiach.\n\n"
    "- **Forma (10%)** — średnia z 5 ostatnich meczów (proxy) względem średniej sezonu "
    "(dodatnia = forma rosnąca, ujemna = spadek).\n\n"
    "- **Dostępność (20%)** — łączne minuty rozegrane w lidze; jak regularnie zawodnik gra "
    "(zaufanie trenera, zdrowie, rola w zespole).\n\n"
    "- **Konsekwencja (15%)** — stabilność ocen meczowych (proxy; mała zmienność = równy poziom).\n\n"
    "Do bazy doliczana jest **premia kontekstowa** (kolumna „Premia”): "
    f"+{B_UP:.2f} za grę ze starszymi rocznikiem, +{B_SEN_SQUAD:.2f} za obecność w kadrze "
    f"seniorów, +{B_SEN_PLAYED:.2f} za realne minuty w seniorach (premie się sumują). "
    "Dlatego PM Index z premią może nieco przekroczyć 1."
)

CSS = """
<style>
.pmrow{display:flex;gap:12px;overflow-x:auto;padding:4px 2px 12px;}
.pmcard{border:1px solid #2b3340;border-radius:12px;padding:12px 14px;width:100%;
        background:#191f29;box-sizing:border-box;}
/* tylko rząd zawierający karty przewija się w bok; kolumny stałej szerokości */
div[data-testid="stHorizontalBlock"]:has(.pmcard){overflow-x:auto !important;flex-wrap:nowrap !important;padding-bottom:6px;}
div[data-testid="stHorizontalBlock"]:has(.pmcard) > div[data-testid="column"],
div[data-testid="stHorizontalBlock"]:has(.pmcard) > div[data-testid="stColumn"]{
        flex:0 0 224px !important;min-width:224px !important;width:224px !important;}
.pmcard h4{margin:0 0 2px;font-size:15px;color:#e8edf4;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.pmcard .sub{font-size:12px;color:#9aa7b6;margin-bottom:8px;line-height:1.4;height:34px;overflow:hidden;}
.pmcard .pm{font-size:22px;font-weight:700;color:#5db0ff;}
.pmcard .pmlbl{font-size:10px;color:#8a97a6;letter-spacing:.5px;text-transform:uppercase;}
.pmcard .row{font-size:12px;color:#c4cdd8;margin-top:6px;}
.pmcard .badges{margin-top:8px;display:flex;flex-wrap:wrap;gap:4px;align-content:flex-start;min-height:72px;}
.b{font-size:10px;padding:2px 7px;border-radius:10px;font-weight:600;white-space:nowrap;}
.b.up{background:#2a2150;color:#c4b5fd;}
.b.sen{background:#16361f;color:#7ee2a0;}
.b.kad{background:#3a2f14;color:#e6c674;}
.b.clj{background:#10303a;color:#6fd6e8;}
/* rozwijana lista filtrow: zawijaj dlugie nazwy zamiast ucinac */
div[data-baseweb="menu"] li{white-space:normal!important;height:auto!important;line-height:1.35;}
div[data-baseweb="popover"] ul{max-height:420px;}
</style>
"""

BADGE_HELP = (
    "**↑ ze starszymi** — zawodnik grał powyżej dominującego rocznika danej ligi, w DOWOLNYCH "
    "rozgrywkach sezonu (np. zawodnik z C1 występujący w A1/B1). Liczba w nawiasie to o ile "
    "roczników wyżej — maksymalny zaobserwowany skok.\n\n"
    "**🪑 kadra seniorów** — był w meczowej kadrze seniorów, ale rozegrał 0 minut.\n\n"
    "**⚽ minuty w seniorach** — rozegrał realne minuty w rozgrywkach seniorskich (na karcie z liczbą minut).\n\n"
    "**🏅 minuty w CLJ** — rozegrał minuty w Centralnej Lidze Juniorów (najwyższy poziom juniorski w PL)."
)


# --------------------------------------------------------------------------- #
def _read_csv(path):
    # jeśli brak pliku .csv, spróbuj spakowanej wersji .csv.gz (pandas wykryje gzip po rozszerzeniu)
    if not os.path.exists(path) and os.path.exists(path + ".gz"):
        path = path + ".gz"
    for enc in ("utf-8", "cp1250", "latin-1"):
        try:
            return pd.read_csv(path, encoding=enc)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return pd.read_csv(path, encoding="latin-1")


def _coerce(df):
    for c in NUMERIC_COMMA:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c].astype(str).str.replace(",", ".", regex=False)
                                  .replace({"NULL": None, "": None, "NaN": None}), errors="coerce")
    for c in NUMERIC_PLAIN:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c].replace({"NULL": None, "": None}), errors="coerce")
    if "match_date" in df.columns:
        df["match_date"] = pd.to_datetime(df["match_date"], errors="coerce")
    for c in ["in_selected_play", "is_junior_comp", "gra_ze_starszymi"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip().str.lower().map({"true": True, "false": False})
    return df


def _pretty_caps(x):
    """Nazwy spoza słownika teamy_kluby bywają KAPITALIKAMI — ładnie je formatujemy,
    zachowując krótkie skróty (MKS, GKS, LKS, KS) w wersji wielkimi literami."""
    if not isinstance(x, str) or any(ch.islower() for ch in x):
        return x
    return " ".join(w if (len(w) <= 3 or not any(c.isalpha() for c in w)) else w.capitalize()
                    for w in x.split())


def _logo_src(logo):
    if isinstance(logo, str) and logo.startswith(("http://", "https://")):
        return logo
    try:
        import base64
        import mimetypes
        mt = mimetypes.guess_type(logo)[0] or "image/png"
        with open(logo, "rb") as fh:
            return f"data:{mt};base64,{base64.b64encode(fh.read()).decode()}"
    except Exception:
        return ""


def _clean_names(df, tk):
    tmap = dict(zip(tk["team_id"], tk["final_team_name"]))
    cmap = dict(zip(tk["club_id"], tk["final_club_name"]))
    if "team_id" in df.columns:
        df["team_name"] = df["team_id"].map(tmap).fillna(df.get("team_name"))
    if "club_id" in df.columns:
        df["club_name"] = df["club_id"].map(cmap).fillna(df.get("club_name"))
    if "opponent_id" in df.columns:
        df["opponent_name"] = df["opponent_id"].map(tmap).fillna(df.get("opponent_name"))
    for col in ("team_name", "club_name", "opponent_name"):
        if col in df.columns:
            df[col] = df[col].map(_pretty_caps)
    return df


@st.cache_data(show_spinner=False)
def load_data(stats_path="stats_test.csv", matches_path="matches_test.csv",
              teamy_path="teamy_kluby_25_26.csv"):
    if DATA_MODE == "db":
        return load_from_db()
    stats, matches = _coerce(_read_csv(stats_path)), _coerce(_read_csv(matches_path))
    if os.path.exists(teamy_path):
        tk = _read_csv(teamy_path)
        stats, matches = _clean_names(stats, tk), _clean_names(matches, tk)
    return stats, matches


def load_from_db(season_id=None, play_id=None):
    import psycopg2
    conn = psycopg2.connect(host=_secret("PGHOST"), dbname=_secret("PGDATABASE"),
                            user=_secret("PGUSER"), password=_secret("PGPASSWORD"),
                            port=_secret("PGPORT", "5432") or "5432")
    p = {"season_id": season_id, "play_id": play_id}
    stats = _coerce(pd.read_sql(open("analiza_stats5_1.sql").read(), conn, params=p))
    matches = _coerce(pd.read_sql(open("analiza_mecze5_2.sql").read(), conn, params=p))
    conn.close()
    return stats, matches


# --------------------------------------------------------------------------- #
def _attrs(stats):
    sel = stats[stats["in_selected_play"] == True].copy()
    sel["zawodnik"] = (sel["firstname"].fillna("") + " " + sel["lastname"].fillna("")).str.strip()
    cols = ["player_id", "zawodnik", "team_name", "club_name", "league_name", "play_name",
            "region_name", "est_birth_year", "rocznik_pewnosc", "rocznik_widelki",
            "status_seniorski", "senior_minutes", "senior_squad_apps"]
    base = sel[[c for c in cols if c in sel.columns]].drop_duplicates("player_id")
    # "Gra ze starszymi" z CAŁEGO sezonu: czy w DOWOLNYCH rozgrywkach zawodnik grał powyżej
    # dominującego rocznika danej ligi (nie tylko w wybranej). To jest sygnał skautingowy —
    # np. zawodnik z C1 występujący w A1/B1 ma znacznik na wierszu A1, nie na wierszu C1.
    up = stats.copy()
    if "gra_ze_starszymi" in up.columns:
        up["_gzs"] = (up["gra_ze_starszymi"].astype(str).str.strip().str.lower()
                      .isin(["true", "1", "t", "yes"]))
    else:
        up["_gzs"] = False
    up["_rwg"] = (pd.to_numeric(up["roczniki_w_gore"], errors="coerce")
                  if "roczniki_w_gore" in up.columns else np.nan)
    agg = (up.groupby("player_id")
           .agg(gra_ze_starszymi=("_gzs", "max"), roczniki_w_gore=("_rwg", "max"))
           .reset_index())
    agg["gra_ze_starszymi"] = agg["gra_ze_starszymi"].fillna(False).astype(bool)
    return base.merge(agg, on="player_id", how="left")


def _play_metrics(g):
    g = g.sort_values("match_date")
    mins, goals = g["minutes"].sum(), g["goals"].sum()
    cards = g["yellow_cards"].sum() + g["red_cards"].sum()
    cidx = g["yellow_cards"].sum() + 2 * g["red_cards"].sum()
    s = g["match_score"].dropna() if "match_score" in g.columns else pd.Series([], dtype=float)
    forma = ((s.tail(5).mean() - s.mean()) / s.mean()) if len(s) >= 3 and s.mean() else np.nan
    return pd.Series({"min_play": mins, "mecze_play": g["match_id"].nunique(), "gole_play": goals,
                      "kartki_play": cards, "score_play": s.mean() if len(s) else np.nan,
                      "gole_per90": (goals / mins * 90) if mins else np.nan,
                      "kartki_per90": (cidx / mins * 90) if mins else np.nan,
                      "konsekwencja": (1 / (1 + s.std(ddof=0))) if len(s) >= 2 else np.nan,
                      "forma": forma})


def _total_metrics(g):
    mins, goals = g["minutes"].sum(), g["goals"].sum()
    s = g["match_score"].dropna() if "match_score" in g.columns else pd.Series([], dtype=float)
    lead = (g.groupby("league_name")["minutes"].sum().idxmax()
            if g["minutes"].sum() > 0 and g["league_name"].notna().any() else None)
    return pd.Series({"min_total": mins, "mecze_total": g["match_id"].nunique(),
                      "gole_total": goals,
                      "kartki_total": g["yellow_cards"].sum() + g["red_cards"].sum(),
                      "score_total": s.mean() if len(s) else np.nan, "liga_wiodaca": lead})


@st.cache_data(show_spinner=False)
def build(_stats, _matches):
    attrs = _attrs(_stats)
    play = _matches[_matches["in_selected_play"] == True]
    base = play.groupby("player_id").apply(_play_metrics).reset_index()
    tot = _matches.groupby("player_id").apply(_total_metrics).reset_index()
    df = attrs.merge(base, on="player_id", how="left").merge(tot, on="player_id", how="left")
    # usuń zawodników bez nazwiska (brak w tabeli players → brak danych, nieprzydatni na liście)
    df = df[df["zawodnik"].fillna("").str.strip() != ""].reset_index(drop=True)

    # --- Realny PlayMaker Score 2.0 per mecz (league-aware przez rank_p dla młodzieży) ---
    # Wyświetlamy pełny score (age_part + stats_part). Ranking liczymy ze stats_part
    # (część league + performance), bo dla młodzieży age_part dominuje i zacierałby poziom ligi;
    # sygnał wieku/gry w górę jest osobno w premiach.
    mall = _matches.copy()
    comp = compute_pm_score(mall)
    mall["_sc"] = comp["score"].values
    mall["_sp"] = comp["stats_part"].values
    mall["_rank_p"] = _rank_p_series(mall)
    mn_all = pd.to_numeric(mall["minutes"], errors="coerce").fillna(0)
    den_t = mn_all.groupby(mall["player_id"]).sum().replace(0, np.nan)
    df["pm_score_total"] = df["player_id"].map((mall["_sc"] * mn_all).groupby(mall["player_id"]).sum() / den_t)

    pl = mall[mall["in_selected_play"] == True]
    mn = mn_all[pl.index]
    den = mn.groupby(pl["player_id"]).sum().replace(0, np.nan)
    df["pm_score"] = df["player_id"].map((pl["_sc"] * mn).groupby(pl["player_id"]).sum() / den)

    # Zakres rankowania jakości: 'selected' = wybrana liga (kluby); 'total' = wszystkie mecze
    # sezonu (raporty "grali w górę" — inaczej zawodnicy grający głównie wyżej mają pustą jakość).
    if QUALITY_SCOPE == "total" or PM_RANK_MODE == "talent":
        qf, qmn, qden = mall, mn_all, den_t
    else:
        qf, qmn, qden = pl, mn, den
    df["pm_quality"] = df["player_id"].map((qf["_sp"] * qmn).groupby(qf["player_id"]).sum() / qden)
    df["rank_p_avg"] = df["player_id"].map((qf["_rank_p"] * qmn).groupby(qf["player_id"]).sum() / qden).round(1)

    # Forma/Konsekwencja ze stats_part per mecz (gęste, league-aware)
    def _form(gp):
        x = gp.sort_values("match_date")["_sp"].dropna()
        m = x.mean()
        return pd.Series({"_forma_px": ((x.tail(5).mean() - m) / m) if len(x) >= 3 and m else np.nan,
                          "_kons_px": (1 / (1 + x.std(ddof=0))) if len(x) >= 2 else np.nan})
    pform = qf.groupby("player_id").apply(_form)
    df["_forma_px"] = df["player_id"].map(pform["_forma_px"])
    df["_kons_px"] = df["player_id"].map(pform["_kons_px"])

    df["Ofensywa"] = df["gole_per90"].rank(pct=True)
    if PM_SCORE_MODE == "v7":
        df["Jakość"] = df["pm_quality"].rank(pct=True)
        df["Forma"] = df["_forma_px"].rank(pct=True)
        df["Konsekwencja"] = df["_kons_px"].rank(pct=True)
    else:
        df["Jakość"] = df["score_play"].rank(pct=True)
        df["Forma"] = df["forma"].rank(pct=True)
        df["Konsekwencja"] = df["konsekwencja"].rank(pct=True)
    df["Dostępność"] = (df["min_total"] if (QUALITY_SCOPE == "total" or PM_RANK_MODE == "talent")
                        else df["min_play"]).rank(pct=True)
    df["Dyscyplina"] = (-df["kartki_per90"]).rank(pct=True)
    # --- "gra ze starszymi": gra w dywizji o STARSZYM max-roczniku niż własny rocznik.
    #     Np. 2011 (własna C2/U-14) w C1/U-15 = +1, w B1/U-17 = +3. Uprawnienie nie chroni —
    #     młodszy zawsze "gra w górę". Seniorzy osobno (znacznik ⚽). ---
    mall["_maxy"] = _cat_maxyear_series(mall)
    _by = df.drop_duplicates("player_id").set_index("player_id")["est_birth_year"]
    py = mall["player_id"].map(_by)                      # rocznik zawodnika (per mecz)
    # "zagrał" = ma minuty LUB dowód występu (gol/kartka) — źródło czasem gubi minuty
    # (np. Oskar: 0 min w C1, ale 2 gole → oczywiście zagrał). Znacznik "w górę" ma to łapać.
    _gg = pd.to_numeric(mall.get("goals"), errors="coerce").fillna(0)
    _yc = pd.to_numeric(mall.get("yellow_cards"), errors="coerce").fillna(0)
    _rc = pd.to_numeric(mall.get("red_cards"), errors="coerce").fillna(0)
    played = (mn_all > 0) | (_gg > 0) | (_yc > 0) | (_rc > 0)
    jun_older = played & mall["_maxy"].notna() & py.notna() & (py > mall["_maxy"])
    nyrs = (py - mall["_maxy"]).where(jun_older)         # o ile roczników wyżej
    rwg = nyrs.groupby(mall["player_id"]).max()
    df["roczniki_w_gore"] = df["player_id"].map(rwg)
    has_jun_older = jun_older.groupby(mall["player_id"]).any()
    df["_jun_older"] = df["player_id"].map(has_jun_older).fillna(False).astype(bool)

    # premia kontekstowa (kolumna „Premia”; w trybie standard wchodzi do PM Index)
    sm = df["senior_minutes"].fillna(0)
    sq = df["senior_squad_apps"].fillna(0)
    # "↑ ze starszymi" = TYLKO starsza dywizja juniorska. Seniorzy -> osobny znacznik ⚽.
    df["gra_ze_starszymi"] = df["_jun_older"].astype(bool)
    df["PM_premia"] = (df["gra_ze_starszymi"].astype(float) * B_UP
                       + (sm > 0).astype(float) * B_SEN_PLAYED
                       + ((sm == 0) & (sq > 0)).astype(float) * B_SEN_SQUAD)

    lg = _matches.groupby("player_id")["league_name"].agg(lambda s: set(s.dropna()))
    pp = _matches.groupby("player_id")["play_name"].agg(lambda s: set(s.dropna()))
    df["_leagues"] = df["player_id"].map(lg)
    df["_plays"] = df["player_id"].map(pp)
    clj = _matches[_matches["league_name"].astype(str)
                   .str.contains(r"\bCLJ\b|Centralna Liga Junior", case=False,
                                 regex=True, na=False)]
    cljm = clj.groupby("player_id")["minutes"].sum()
    df["clj_minutes"] = df["player_id"].map(cljm).fillna(0)

    # minuty w futsalu / halówce (do znacznika "halowiec")
    _fut = _matches[
        _matches["league_name"].astype(str).str.contains(r"futsal|PLF|halow", case=False, regex=True, na=False)
        | _matches["play_name"].astype(str).str.contains(r"futsal|halow", case=False, regex=True, na=False)
    ]
    df["futsal_minutes"] = df["player_id"].map(_fut.groupby("player_id")["minutes"].sum()).fillna(0)

    # minuty zagrane "w górę" (w starszej dywizji) — sygnał poziomu w trybie talent.
    # Skok ważony liczbą roczników, ale z CZAPKĄ (PM_SKOK_CAP, domyślnie 2): granie
    # +3/+4 (często dziecko wstawione z braku kadry, nie z klasy) nie rakietuje rankingu.
    _cap = float(_secret("PM_SKOK_CAP", "2") or 2)
    _yrs = (py - mall["_maxy"]).clip(upper=_cap)
    up_w = (mn_all * _yrs).where(jun_older, 0)
    up2 = up_w.groupby(mall["player_id"]).sum()
    df["up2_min"] = df["player_id"].map(up2).fillna(0)

    if PM_RANK_MODE == "talent":
        # mix: jakość (league-aware) + poziom (CLJ/seniorzy/skok) + wolumen
        Q = df["pm_quality"].rank(pct=True)
        lvl_raw = (df["clj_minutes"].fillna(0) * W_CLJ
                   + df["senior_minutes"].fillna(0) * W_SENIOR
                   + df["up2_min"].fillna(0) * W_SKOK)
        df["Poziom"] = lvl_raw.rank(pct=True)
        Vol = df["min_total"].fillna(0).rank(pct=True)
        wsum = (W_JAKOSC + W_POZIOM + W_WOLUMEN) or 1.0
        df["PM_base"] = (W_JAKOSC * Q + W_POZIOM * df["Poziom"] + W_WOLUMEN * Vol) / wsum
        # poziom/wolumen już zawierają sygnał "w górę" → bez podwójnego liczenia premii
        df["PM_Index"] = df["PM_base"]
    else:
        df["PM_base"] = sum(df[a].fillna(0) * w for a, w in PM_WEIGHTS.items())
        df["PM_Index"] = df["PM_base"] + df["PM_premia"]

    ry = _matches["match_date"].dt.year.max()
    df["_ref_year"] = int(ry) if pd.notna(ry) else 2026

    # lokalizacja / dystans do Opola (opcjonalne, z zawodnicy_lokalizacja.csv)
    try:
        _loc = pd.read_csv("zawodnicy_lokalizacja.csv", dtype=str, keep_default_na=False)
        _loc["player_id"] = _loc["player_id"].astype(str)
        if "miejscowosc" in _loc.columns:
            df["miejscowosc"] = df["player_id"].map(dict(zip(_loc["player_id"], _loc["miejscowosc"])))
        if "spoza_regionu" in _loc.columns:
            df["spoza_regionu"] = df["player_id"].map(dict(zip(_loc["player_id"], _loc["spoza_regionu"])))
        for c in ("lat", "lon", "km_do_opola"):
            if c in _loc.columns:
                df[c] = df["player_id"].map(dict(zip(_loc["player_id"],
                                                      pd.to_numeric(_loc[c], errors="coerce"))))
    except Exception:
        pass
    return df


def badges_html(r):
    out = []
    if bool(r.get("gra_ze_starszymi")):
        n = r.get("roczniki_w_gore")
        lbl = (f"↑ ze starszymi (+{int(n)})" if pd.notna(n) and n and n >= 1
               else "↑ ze starszymi")
        out.append(f'<span class="b up">{lbl}</span>')
    sm = r.get("senior_minutes") or 0
    if sm > 0:
        out.append(f'<span class="b sen">⚽ {int(sm)}′ w seniorach</span>')
    elif (r.get("senior_squad_apps") or 0) > 0:
        out.append('<span class="b kad">🪑 kadra seniorów</span>')
    cm = r.get("clj_minutes") or 0
    if cm > 0:
        out.append(f'<span class="b clj">🏅 {int(cm)}′ w CLJ</span>')
    return "".join(out)


EXPORT_COLS = [("zawodnik", L["player_one"]), ("club_name", "Klub"), ("team_name", "Drużyna"),
               ("region_name", "Województwo"), ("est_birth_year", "Rocznik"),
               ("liga_wiodaca", "Liga wiodąca"),
               ("PM_Index", "PM Index"), ("PM_premia", "Premia"),
               ("pm_score", "Score (liga)"), ("pm_score_total", "Score (total)"),
               ("rank_p_avg", "Poziom"),
               ("min_play", "Min (liga)"),
               ("mecze_play", "Mecze (liga)"), ("gole_play", "Gole (liga)"),
               ("min_total", "Min (total)"), ("mecze_total", "Mecze (total)"),
               ("senior_minutes", "Min seniorzy"), ("clj_minutes", "Min CLJ"),
               ("gra_ze_starszymi", "Ze starszymi"),
               ("status_seniorski", "Status senior")]


def export_frame(f, top_n, per_region=False):
    cols = [(s, d) for s, d in EXPORT_COLS if s in f.columns]
    f = f.sort_values("PM_Index", ascending=False)
    if per_region and "region_name" in f.columns:
        # top N z KAŻDEGO województwa osobno; Lp. numeruje się od nowa w każdym woj.
        base = (f.groupby("region_name", sort=True, group_keys=False)
                 .head(int(top_n))
                 .sort_values(["region_name", "PM_Index"], ascending=[True, False]))
        lp = list(base.groupby("region_name").cumcount() + 1)
    else:
        base = f.head(int(top_n))
        lp = list(range(1, len(base) + 1))
    df = base[[s for s, _ in cols]].rename(columns=dict(cols)).copy()
    for c in ["PM Index", "Premia", "Score (liga)"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(3)
    df.insert(0, "Lp.", lp)
    return df


def build_excel(df, title):
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Zestawienie", startrow=2)
        ws = xw.sheets["Zestawienie"]
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=13)
        for ci in range(1, df.shape[1] + 1):
            cell = ws.cell(row=3, column=ci)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E5AAC")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A4"
        for i, col in enumerate(df.columns, start=1):
            ml = max([len(str(col))] + [len(str(v)) for v in df[col].head(300)])
            ws.column_dimensions[get_column_letter(i)].width = min(max(ml + 2, 8), 38)
    return buf.getvalue()


def _i(x):
    """int odporny na NaN/None (NaN or 0 zwraca NaN, a int(NaN) rzuca ValueError)."""
    try:
        v = float(x)
        return 0 if pd.isna(v) else int(v)
    except (TypeError, ValueError):
        return 0


def _card_html(r):
    by = r.get("est_birth_year")
    age = f"{int(r['_ref_year'] - by)} lat" if pd.notna(by) else "—"
    rok = f"rocznik {int(by)}" if pd.notna(by) else ""
    klub = r.get("club_name") or r.get("team_name") or "—"
    lead = r.get("liga_wiodaca") or r.get("league_name") or "—"
    return (f'<div class="pmcard"><h4>{r["zawodnik"]}</h4>'
            f'<div class="sub">{klub}<br>liga wiodąca: {lead}</div>'
            f'<div class="pmlbl">PM Index</div><div class="pm">{r["PM_Index"]:.2f}</div>'
            f'<div class="row">{rok} · {age}</div>'
            f'<div class="row">{_i(r.get("min_total"))} min · {_i(r.get("mecze_total"))} meczów (sezon)</div>'
            f'<div class="badges">{badges_html(r)}</div></div>')


# --------------------------------------------------------------------------- #
# --------------------------------------------------------------------------- #
def check_password():
    """Lekka bramka: jesli ustawiono APP_PASSWORD (st.secrets lub env) - wymagaj.
    Brak hasla = otwarte (lokalnie)."""
    pw = _secret("APP_PASSWORD")
    if not pw:
        return True
    if st.session_state.get("auth_ok"):
        return True
    st.title("Almanach ligowy")
    with st.form("login"):
        x = st.text_input("Hasło dostępu", type="password")
        if st.form_submit_button("Wejdź"):
            if x == str(pw):
                st.session_state["auth_ok"] = True
                st.rerun()
            else:
                st.error("Błędne hasło.")
    return False


def _visits_file():
    return _secret("PM_VISITS_FILE", "") or "visits.json"


def _bump_visits():
    """Zlicza otwarcia (raz na sesję). Zapis do pliku JSON: {total, days:{YYYY-MM-DD:n}, last}.
    Wszystko w try/except — błąd licznika NIGDY nie może wywalić aplikacji."""
    try:
        if st.session_state.get("_counted"):
            return st.session_state.get("_visit_total")
        st.session_state["_counted"] = True
        import datetime
        path = _visits_file()
        data = {"total": 0, "days": {}}
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as fh:
                    data = json.load(fh)
            except Exception:
                data = {"total": 0, "days": {}}
        today = datetime.date.today().isoformat()
        data["total"] = int(data.get("total", 0)) + 1
        data.setdefault("days", {})
        data["days"][today] = int(data["days"].get(today, 0)) + 1
        data["last"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False)
        st.session_state["_visit_total"] = data["total"]
        return data["total"]
    except Exception:
        return None


def _show_visits():
    """Panel statystyk — widoczny tylko pod adresem z ?stats=1 (dla Ciebie, nie dla klubu)."""
    try:
        if str(st.query_params.get("stats", "")) not in ("1", "true", "tak"):
            return
        path = _visits_file()
        if not os.path.exists(path):
            st.info("Licznik jeszcze pusty (brak zapisanych otwarć).")
            return
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        st.markdown("### 📈 Statystyki otwarć")
        c = st.columns(3)
        c[0].metric("Otwarcia łącznie", int(data.get("total", 0)))
        days = data.get("days", {})
        import datetime
        last7 = sum(v for d, v in days.items()
                    if d >= (datetime.date.today() - datetime.timedelta(days=6)).isoformat())
        c[1].metric("Ostatnie 7 dni", last7)
        c[2].metric("Ostatnie otwarcie", data.get("last", "—"))
        if days:
            srt = pd.DataFrame(sorted(days.items()), columns=["Dzień", "Otwarcia"]).set_index("Dzień")
            st.bar_chart(srt.tail(30))
        st.caption("Uwaga: na darmowym Streamlit Cloud licznik zeruje się po restarcie/uśpieniu "
                   "aplikacji (plik jest tymczasowy). To dolna granica od ostatniego startu.")
    except Exception:
        pass


def _intro_md():
    """Tekst wprowadzający raportu. Źródło: sekret PM_INTRO albo plik intro.md w repo.
    Brak obu → nic się nie pokazuje (kluby bez intro.md są nietknięte)."""
    txt = _secret("PM_INTRO", "")
    if not txt and os.path.exists("intro.md"):
        try:
            with open("intro.md", encoding="utf-8") as fh:
                txt = fh.read()
        except Exception:
            txt = ""
    return txt


def main():
    st.set_page_config(page_title="Almanach ligowy", layout="wide")
    st.markdown(CSS, unsafe_allow_html=True)
    if not check_password():
        st.stop()
    _bump_visits()
    _show_visits()
    stats, matches = load_data()
    data = build(stats, matches)

    sel_mask = stats["in_selected_play"] == True
    sel = stats[sel_mask]
    plays = sorted(sel["play_name"].dropna().unique())
    leagues = sorted(sel["league_name"].dropna().unique()) if "league_name" in sel.columns else []
    regions = sorted(r for r in (sel["region_name"].dropna().unique()
                                 if "region_name" in stats.columns else []) if r)
    if len(plays) <= 1:
        liga = plays[0] if plays else "—"
    elif len(leagues) == 1:
        liga = f"{leagues[0]} — {len(plays)} rozgrywek"
    else:
        liga = f"{len(plays)} rozgrywek ({', '.join(leagues[:4])}{'…' if len(leagues) > 4 else ''})"
    region_txt = ", ".join(regions) or REGION
    logo = _secret("CLUB_LOGO_URL", "")
    if not logo:
        for _p in ("logo.png", "logo.jpg", "logo.jpeg", "logo.webp"):
            if os.path.exists(_p):
                logo = _p
                break

    def _header_text():
        st.title("Almanach ligowy")
        reg = f"Region: **{region_txt}**  ·  " if region_txt else ""
        st.markdown(f"{reg}**{liga}** · {L['players_gen']}: {len(data)} · "
                    f"z minutami w seniorach: {(data['senior_minutes'].fillna(0) > 0).sum()}")
        if len(plays) > 1:
            st.caption(f"Zestawienie obejmuje {len(plays)} rozgrywek (cały wybrany poziom / region). "
                       "Kolumny „(liga)” = łącznie z całego wybranego zakresu, „(total)” = cały sezon "
                       f"{L['of_player']} we wszystkich rozgrywkach.")
        else:
            st.caption(f"{L['players_all']} z wybranej ligi wraz z meczami w bieżącym sezonie we "
                       "wszystkich rozgrywkach. Kolumny „(liga)” dotyczą wybranej ligi, „(total)” — "
                       "całego sezonu.")

    src = _logo_src(logo) if logo else ""
    if src:
        hc = st.columns([6, 1])
        with hc[0]:
            _header_text()
        hc[1].markdown(
            '<div style="display:flex;justify-content:flex-end;align-items:flex-start;">'
            f'<img src="{src}" style="width:auto;max-width:100%;max-height:104px;'
            'object-fit:contain;">'
            "</div>", unsafe_allow_html=True)
    else:
        _header_text()

    _intro = _intro_md()
    if _intro:
        with st.expander("ℹ️ O tym raporcie / jak korzystać", expanded=True):
            st.markdown(_intro)

    # ---- FILTRY ----
    # Reset przez "nonce": po wyczyszczeniu zmieniamy klucze widżetów, więc wszystkie
    # (także suwaki) startują od wartości domyślnych — to pewniejsze niż pop+rerun.
    _fn = st.session_state.get("_filter_nonce", 0)
    def K(base):
        return f"{base}_{_fn}"
    with st.container(border=True):
        ch = st.columns([6, 1])
        ch[0].markdown("**Filtry**")
        if ch[1].button("🧹 Wyczyść filtry", use_container_width=True, type="primary"):
            st.session_state["_filter_nonce"] = _fn + 1
            st.rerun()
        has_region = "region_name" in data.columns and data["region_name"].notna().any()
        if has_region:
            rr = st.columns([2, 6])
            f_reg = rr[0].multiselect("Województwo",
                                      sorted(data["region_name"].dropna().unique()), key=K("f_woj"))
        else:
            f_reg = []
        r1 = st.columns([2, 2, 2])
        q = r1[0].text_input(f"{L['player_one']} (imię/nazwisko)", "", key=K("f_zaw"))
        f_club = r1[1].multiselect("Klub", sorted(data["club_name"].dropna().unique()), key=K("f_klub"))
        f_lg = r1[2].multiselect(f"Rozgrywki (gdziekolwiek {L['played']})",
                                 sorted({x for s in data["_leagues"].dropna() for x in s}), key=K("f_rozgr"))
        f_pl = st.multiselect(f"Liga (gdziekolwiek {L['played']})",
                              sorted({x for s in data["_plays"].dropna() for x in s}), key=K("f_liga"))
        r2 = st.columns(4)
        def rng(col, label, c, key, integer=False):
            lo, hi = float(np.nanmin(data[col])), float(np.nanmax(data[col]))
            if not np.isfinite(lo) or lo == hi:
                return (lo, hi)
            if integer:
                lo, hi = int(np.floor(lo)), int(np.ceil(hi))
                if lo == hi:
                    return (lo, hi)
                return c.slider(label, lo, hi, (lo, hi), step=1, key=key)
            return c.slider(label, lo, hi, (lo, hi), key=key)
        s_score = rng("pm_score", "Score (liga)", r2[0], K("f_score"))
        s_min = rng("min_play", "Minuty (liga)", r2[1], K("f_min"), integer=True)
        s_mecz = rng("mecze_play", "Mecze (liga)", r2[2], K("f_mecz"), integer=True)
        s_kart = rng("kartki_total", "Kartki total", r2[3], K("f_kart"), integer=True)
        r3 = st.columns(4)
        f_up = r3[0].checkbox("↑ Gra ze starszymi", key=K("f_up"))
        f_kad = r3[1].checkbox("🪑 W kadrze seniorów", key=K("f_kad"))
        f_sen = r3[2].checkbox("⚽ Minuty w seniorach", key=K("f_sen"))
        f_clj = r3[3].selectbox("CLJ", ["— wszyscy —", "🏅 tylko z CLJ", "🚫 bez CLJ (nie grał)"],
                                key=K("f_clj"))
        gole_prog = int(float(_secret("PM_GOLE_PROG", "50") or 50))
        pluca_prog = int(float(_secret("PM_PLUCA_PROG", "5000") or 5000))
        r4 = st.columns(4)
        f_gole = r4[0].checkbox(f"🎯 >{gole_prog} goli", key=K("f_gole"))
        f_pluca = r4[1].checkbox(f"🫁 Żelazne płuca (>{pluca_prog}')", key=K("f_pluca"))
        f_futsal = r4[2].checkbox("🥅 Halowiec (futsal)", key=K("f_futsal"))
        f_skok2 = r4[3].checkbox("↑↑ Skok 2+ roczniki", key=K("f_skok2"))
        km_max = 0
        if "km_do_opola" in data.columns and data["km_do_opola"].notna().any():
            _hi = int(pd.to_numeric(data["km_do_opola"], errors="coerce").max() or 0)
            km_max = st.slider("📍 Maks. odległość do Opola (km, 0 = bez filtra)",
                               0, _hi, 0, step=5, key=K("f_km"))

    # ---- FILTROWANIE ----
    f = data.copy()
    f = f.sort_values("PM_Index", ascending=False).reset_index(drop=True)
    f["Lp"] = range(1, len(f) + 1)          # globalne miejsce w rankingu (przed filtrami)
    if f_reg:
        f = f[f["region_name"].isin(f_reg)]
    if q:
        f = f[f["zawodnik"].str.contains(q, case=False, na=False)]
    if f_club:
        f = f[f["club_name"].isin(f_club)]
    if f_lg:
        f = f[f["_leagues"].apply(lambda s: bool(s & set(f_lg)) if isinstance(s, set) else False)]
    if f_pl:
        f = f[f["_plays"].apply(lambda s: bool(s & set(f_pl)) if isinstance(s, set) else False)]
    for col, (lo, hi) in [("pm_score", s_score), ("min_play", s_min),
                          ("mecze_play", s_mecz), ("kartki_total", s_kart)]:
        f = f[f[col].fillna(-1).between(lo, hi) | f[col].isna()]
    if f_up:
        f = f[f["gra_ze_starszymi"].fillna(False).astype(bool)]
    if f_kad:
        f = f[(f["senior_squad_apps"].fillna(0) > 0) & (f["senior_minutes"].fillna(0) == 0)]
    if f_sen:
        f = f[f["senior_minutes"].fillna(0) > 0]
    if f_clj == "🏅 tylko z CLJ":
        f = f[f["clj_minutes"].fillna(0) > 0]
    elif isinstance(f_clj, str) and f_clj.startswith("🚫"):
        f = f[f["clj_minutes"].fillna(0) == 0]
    if f_gole:
        f = f[f["gole_total"].fillna(0) > gole_prog]
    if f_pluca:
        f = f[f["min_total"].fillna(0) > pluca_prog]
    if f_futsal:
        f = f[f["futsal_minutes"].fillna(0) > 0]
    if f_skok2:
        f = f[f["roczniki_w_gore"].fillna(0) >= 2]
    if km_max and "km_do_opola" in f.columns:
        f = f[f["km_do_opola"].notna() & (f["km_do_opola"] <= km_max)]
    f = f.sort_values("PM_Index", ascending=False).reset_index(drop=True)

    # ---- KARTY TOPOWYCH (jeden rząd, przewijany w bok; natywny wybór — lekki rerun) ----
    st.markdown(f"### 🏅 {L['top_players']}")
    if len(f):
        topcards = f.head(15).reset_index(drop=True)
        cols = st.columns(len(topcards))
        for j in range(len(topcards)):
            r = topcards.iloc[j]
            with cols[j]:
                st.markdown(_card_html(r), unsafe_allow_html=True)
                if st.button("Wybierz", key=f"card_{r['player_id']}",
                             use_container_width=True):
                    st.session_state["sel_pid"] = r["player_id"]
                    st.rerun()
    else:
        st.info(L["no_players"])

    # ---- TABELA ----
    st.markdown("### 📋 Analityka")
    ci = st.columns([2, 2, 2, 2, 3])
    with ci[0].popover("ℹ️ Czym jest PM Index?"):
        st.markdown(PM_HELP)
    with ci[1].popover("🏷️ Znaczniki"):
        st.markdown(BADGE_HELP)
    max_n = max(10, len(f))
    top_n = ci[2].number_input("Top N", min_value=10, max_value=max_n,
                               value=min(100, max_n), step=10)
    per_woj = ci[3].checkbox("na każde województwo", key=K("f_perwoj"),
                             help="Zamiast top N ogółem — top N z KAŻDEGO województwa osobno "
                                  "(np. top 40 z każdego woj., połączone w jeden plik).")
    exp = export_frame(f, top_n, per_region=per_woj)
    title = f"Almanach ligowy — {liga}" + (f" — {region_txt}" if region_txt else "")
    if per_woj:
        title += f"  ·  top {int(top_n)} / województwo"
    xlsx = build_excel(exp, title)
    if xlsx is not None:
        ci[4].download_button("⬇️ Pobierz zestawienie (Excel)", xlsx,
                              file_name="zestawienie_playmaker.xlsx",
                              mime=("application/vnd.openxmlformats-officedocument."
                                    "spreadsheetml.sheet"),
                              use_container_width=True)
    else:
        ci[4].download_button("⬇️ Pobierz zestawienie (CSV)",
                              exp.to_csv(index=False).encode("utf-8-sig"),
                              file_name="zestawienie_playmaker.csv", mime="text/csv",
                              use_container_width=True)

    def znaczniki(r):
        z = []
        if bool(r.get("gra_ze_starszymi")): z.append("↑")
        if (r.get("senior_minutes") or 0) > 0: z.append("⚽")
        elif (r.get("senior_squad_apps") or 0) > 0: z.append("🪑")
        if (r.get("clj_minutes") or 0) > 0: z.append("🏅")
        return " ".join(z)
    # wybór z karty ma priorytet (przycisk ustawia st.session_state["sel_pid"] — lekki rerun)
    sel_card = st.session_state.get("sel_pid")
    sel_card = sel_card if (sel_card and (f["player_id"] == sel_card).any()) else None

    ft = f.copy()
    ft["Znaczniki"] = ft.apply(znaczniki, axis=1)
    cmap = {"Lp": "#", "zawodnik": L["player_one"], "Znaczniki": "Znaczniki", "region_name": "Województwo",
            "team_name": "Drużyna",
            "club_name": "Klub", "est_birth_year": "Rocznik", "rocznik_pewnosc": "Pewność",
            "miejscowosc": "Miejscowość", "km_do_opola": "~km do Opola", "PM_Index": "PM Index",
            "PM_premia": "Premia", "pm_score": "Score (liga)", "pm_score_total": "Score (total)",
            "rank_p_avg": "Poziom",
            "min_play": "Min (liga)", "min_total": "Min (total)",
            "mecze_play": "Mecze (liga)", "mecze_total": "Mecze (total)",
            "gole_play": "Gole (liga)", "gole_total": "Gole (total)",
            "kartki_total": "Kartki", "senior_minutes": "Min. seniorzy",
            "clj_minutes": "Min. CLJ"}

    if sel_card:
        who = f.loc[f["player_id"] == sel_card, "zawodnik"].iloc[0]
        cc = st.columns([5, 2, 2])
        cc[0].success(f"Wybrany zawodnik: **{who}** — analityka i mecze zawężone do niego.")
        if cc[1].button("← Pokaż wszystkich", use_container_width=True):
            st.session_state.pop("sel_pid", None)
            st.rerun()
        _prow = f.loc[f["player_id"] == sel_card].iloc[0]
        _pdf = _zaproszenie_pdf(who, str(_prow.get("est_birth_year", "") or "").split(".")[0],
                                str(_prow.get("club_name", "") or ""))
        if _pdf:
            cc[2].download_button("📄 Zaproszenie na testy", _pdf,
                                  file_name=f"zaproszenie_{who.replace(' ', '_')}.pdf",
                                  mime="application/pdf", use_container_width=True)
        ftab = ft[ft["player_id"] == sel_card]
        sel_pid = sel_card
        select_mode = "ignore"
    else:
        ftab = ft
        sel_pid = None
        select_mode = "rerun"

    disp = ftab[[c for c in cmap if c in ftab.columns]].rename(columns=cmap)
    event = st.dataframe(
        disp, use_container_width=True, height=285, hide_index=True,
        on_select=select_mode, selection_mode="single-row",
        column_config={
            "PM Index": st.column_config.NumberColumn(format="%.2f", help=PM_HELP),
            "Score (liga)": st.column_config.NumberColumn(format="%.3f",
                help="Realny PlayMaker Score 2.0 (v7) — średnia ważona minutami z meczów w wybranym "
                     "zakresie. Dla młodzieży leagueMultiplier z poziomu rozgrywek (rank_p)."),
            "Score (total)": st.column_config.NumberColumn(format="%.3f",
                help="Realny PlayMaker Score 2.0 (v7) — średnia ważona minutami ze wszystkich meczów sezonu."),
            "Poziom": st.column_config.NumberColumn(format="%.1f",
                help="Średni rank_p rozgrywek (0 = topowe juniorskie / CLJ, 10 = słabe lokalne). Niżej = mocniej."),
            "Premia": st.column_config.NumberColumn(format="%.2f",
                help="Premia kontekstowa doliczona do PM Index (gra ze starszymi / kadra / minuty w seniorach)."),
            "Znaczniki": st.column_config.TextColumn(
                help="↑ gra ze starszymi · 🪑 w kadrze seniorów · ⚽ minuty w seniorach · 🏅 minuty w CLJ")})

    if sel_pid is None and event.selection.rows:
        sel_pid = ftab.iloc[event.selection.rows[0]]["player_id"]

    # ---- PODSUMOWANIE SEZONU (per rozgrywki) — zawsze widoczne, jak mecze ----
    if sel_pid:
        who_r = f.loc[f["player_id"] == sel_pid, "zawodnik"].iloc[0]
        st.markdown(f"### 📊 Podsumowanie sezonu: {who_r} — mecze / minuty / gole per liga")
        pv = matches[matches["player_id"] == sel_pid].copy()
    else:
        st.markdown(f"### 📊 Podsumowanie sezonu ({len(f)} {L['players_gen']}) — "
                    f"{L['click_one']} wyżej, by zawęzić")
        pv = matches[matches["player_id"].isin(f["player_id"])].copy()
    pv["_rozg"] = pv["play_name"].map(_rozgrywki_key)
    pv["_min"] = pd.to_numeric(pv["minutes"], errors="coerce")
    pv["_gol"] = pd.to_numeric(pv["goals"], errors="coerce")
    _gcols = (["player_id"] if not sel_pid else []) + ["_rozg", "league_name"]
    agg = (pv.groupby(_gcols)
             .agg(Mecze=("match_id", "nunique"), Min=("_min", "sum"), Gole=("_gol", "sum"))
             .reset_index())
    agg["Gole/90"] = (agg["Gole"] / agg["Min"].replace(0, np.nan) * 90).round(2)
    for c in ("Min", "Gole"):
        agg[c] = agg[c].fillna(0).astype(int)
    agg = agg.rename(columns={"_rozg": "Rozgrywki", "league_name": "Liga"})
    _lr = [_liga_rank(l, r) for l, r in zip(agg["Liga"], agg["Rozgrywki"])]
    agg["_lr0"] = [x[0] for x in _lr]
    agg["_lr1"] = [x[1] for x in _lr]
    if not sel_pid:
        agg["Zawodnik"] = agg["player_id"].map(f.set_index("player_id")["zawodnik"])
        agg = agg.sort_values(["Zawodnik", "_lr0", "_lr1", "Min"],
                              ascending=[True, True, True, False])
        _scols = ["Zawodnik", "Rozgrywki", "Liga", "Mecze", "Min", "Gole", "Gole/90"]
    else:
        agg = agg.sort_values(["_lr0", "_lr1", "Min"], ascending=[True, True, False])
        _scols = ["Rozgrywki", "Liga", "Mecze", "Min", "Gole", "Gole/90"]
    st.dataframe(agg[_scols], use_container_width=True, height=215, hide_index=True,
                 column_config={"Gole/90": st.column_config.NumberColumn(format="%.2f")})

    # ---- MECZE ----
    if sel_pid:
        who = f.loc[f["player_id"] == sel_pid, "zawodnik"].iloc[0]
        st.markdown(f"### ⚽ Mecze: {who}")
        mm = matches[matches["player_id"] == sel_pid]
    else:
        st.markdown(f"### ⚽ Mecze ({len(f)} {L['players_gen']}) — {L['click_one']} wyżej, by zawęzić")
        mm = matches[matches["player_id"].isin(f["player_id"])]
    mm = mm.assign(pm_score=compute_pm_score(mm)["score"].values)
    mc = {"match_date": "Data", "region_name": "Województwo", "league_name": "Liga",
          "play_name": "Play",
          "team_name": "Drużyna", "opponent_name": "Przeciwnik", "team_side": "Strona",
          "match_result": "Wynik", "minutes": "Min", "goals": "Gole",
          "yellow_cards": "ŻK", "red_cards": "CK", "pm_score": "Score",
          "status_seniorski": "Status senior"}
    mshow = (mm.sort_values("match_date", ascending=False)
               [[c for c in mc if c in mm.columns]].rename(columns=mc))
    st.dataframe(mshow, use_container_width=True, height=285, hide_index=True,
                 column_config={"Score": st.column_config.NumberColumn(format="%.3f")})

    # ---- MAPA: odległość do Opola ----
    if "lat" in f.columns and f["lat"].notna().any():
        st.markdown("### 🗺️ Mapa — odległość do Opola")
        mp = f[f["lat"].notna() & f["lon"].notna()].copy()
        if "spoza_regionu" not in mp.columns:
            mp["spoza_regionu"] = False
        mp["spoza_regionu"] = mp["spoza_regionu"].astype(str).str.lower().isin(["true", "1", "tak"])
        # skupiska: 1 punkt na miejscowość, promień ~ liczba zawodników
        grp = (mp.groupby(["miejscowosc", "lat", "lon"], dropna=False)
                 .agg(zawodnikow=("player_id", "nunique"),
                      km=("km_do_opola", "min"),
                      spoza=("spoza_regionu", "max"))
                 .reset_index())
        n_sp = int(grp["spoza"].sum())
        st.caption(f"{int(grp['zawodnikow'].sum())} zawodników w {len(grp)} miejscowościach. "
                   f"🔵 w regionie · 🟠 spoza regionu ({n_sp}) · 🔴 Opole (cel). "
                   f"Wielkość kropki = liczba zawodników.")
        try:
            import pydeck as pdk
            grp["radius"] = 2000 + grp["zawodnikow"] ** 0.5 * 2200
            grp["color"] = grp["spoza"].map(lambda s: [255, 140, 0, 180] if s else [31, 119, 180, 170])
            opole = pd.DataFrame([{"lat": 50.6751, "lon": 17.9213, "miejscowosc": "OPOLE (cel)",
                                   "zawodnikow": 0, "radius": 4000}])
            l_pts = pdk.Layer("ScatterplotLayer", data=grp, get_position="[lon, lat]",
                              get_fill_color="color", get_radius="radius", pickable=True)
            l_op = pdk.Layer("ScatterplotLayer", data=opole, get_position="[lon, lat]",
                             get_fill_color="[214, 39, 40, 230]", get_radius="radius")
            st.pydeck_chart(pdk.Deck(
                layers=[l_pts, l_op],
                initial_view_state=pdk.ViewState(latitude=50.67, longitude=17.92, zoom=7.2),
                tooltip={"text": "{miejscowosc}\n{zawodnikow} zawodników\n~{km} km do Opola"},
                map_style=None))
        except Exception:
            st.map(grp.rename(columns={"lat": "latitude", "lon": "longitude"})[["latitude", "longitude"]])
        with st.expander("📍 Skupiska zawodników wg miejscowości"):
            _t = grp.sort_values("zawodnikow", ascending=False).rename(
                columns={"miejscowosc": "Miejscowość", "zawodnikow": "Zawodników", "km": "~km do Opola"})
            _t["Region"] = _t["spoza"].map(lambda s: "spoza" if s else "opolskie")
            st.dataframe(_t[["Miejscowość", "Zawodników", "~km do Opola", "Region"]],
                         use_container_width=True, hide_index=True, height=250)


if __name__ == "__main__":
    main()